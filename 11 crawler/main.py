import requests
from bs4 import BeautifulSoup as bs
from konlpy.tag import Okt
from openpyxl import load_workbook as load # 엑셀파일 다루기위해

URL = "https://academy.nomadcoders.co/p/instaclone-2-0"
SAVE_DIR = "c:/Users/last2018/dev/python/basic/11 crawler/rank.xlsx" #엑셀파일위치


# http 상태코드를 통해 정상적으로 처리가 됐다면 작동하도록 만들자
def get_tags() : 
    tags_list = [] # 결과값을 리스트객체로 만들자
    rq = requests.get(URL)
    print(rq.status_code) # http상태코드 출력
    if rq.status_code == 200 : # 정상처리됐다면
        soup = bs(rq.content,'html.parser') 
        
        # 리스트형태, 객체형태로 반환되기에 li_list에 할당함
        li_list = soup.find_all('div',class_= 'col-sm-12 course-section') 
        for li in li_list :
            ul = li.select('div > ul') # li도 BeautifulSoup의 메소드사용가능
            # ul객체도 li하나하나를 가지는 리스트객체 이므로 for문 사용
            for l in ul : 
                tags_list.append(l.get_text()) # 결과값을 리스트객체로 만듦
        return tags_list
    else : 
        raise Exception('응답 코드가 다릅니다')


# 엑셀파일에 저장하는 함수
def save_excel(rank) : # rank호출해서 처리함
    wb = load(SAVE_DIR)
    ws = wb.create_sheet('rank')
    idx = 1 # A행에 쓸 인덱스
    # 외부자원사용하기에 에러발생시 대처해야함
    try : 
        for k, v in rank[:15] :
            # A행의 셀에 저장, 문자열(인덱스)
            ws['A' + str(idx)] = "{}({})".format(k,v)
            idx += 1
        wb.save(SAVE_DIR)
    except Exception as e : 
        print(e)
    finally : 
        wb.close()

# 함수호출 > tags_list반환 > 데이터 처리 함수 작성
# 불러온 목록에서 명사를 추출
# 추출한 명사 중 가장 많이 나온 명사 랭킹화 출력
def get_rank() : 
    tags_list = get_tags() # 함수를 통해 문자값을 넘겨받는 객체
    ok = Okt()
    rank = {} # 랭크매길 딕셔너리, 명사값처리
    for tag in tags_list : 
        noun = ok.nouns(tag) # 명사를 따로 구분지음. 고로 리스트
        for n in noun : # noun도 리스트이기에 반복문으로 처리가능
            if not (n in rank) : # 랭크dic에 명사값이 없다면
                rank[n] = 0 # n할당하여 초기화
            rank[n] += 1 # 있다면 랭크값에 1씩 더해줌
    
    # 랭크구하기
    # 내림차순
    rank = sorted(rank.items(), key = lambda x:x[1], reverse = True)
    save_excel(rank)
    # 정렬 결과, 키 값 형태로 포맷
    # for k,v in rank[1:15]: # 1위 ~ 15위까지 한 줄
    #     print("{}({})".format(k,v), 둥 = ' ')

# get_rank()함수 호출
if __name__ == "__main__" : 
    get_rank()

