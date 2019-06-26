from openpyxl import load_workbook as load

DIR = 'C:/Users/last2018/dev/python/basic/9 Excel/test.xlsx'
wb = load(DIR)

# ws = wb.create_sheet('test') # 엑셀파일의 'test'란 이름의 새시트를 만듦
# ws['A1'] = "title1"
# ws['B1'] = "title2"
# wb.save(DIR)
# wb.close()

# 앞에서 작성한 test시트를 가져와서 사용
ws = wb['test']
a2 = ws['A2'].value # 이 A2시트값을 가져옴
b2 = ws['B2'].value
print(a2, b2) # A2,B2에 있는 값을 출력

wb.close() # 객체(외부자원)를 모두 사용하고나면 close해줘야 된다. (외부자원,엑셀,DB)